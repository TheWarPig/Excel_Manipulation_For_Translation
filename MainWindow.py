import os
import re
import tkinter as tk
from tkinter import filedialog
import openpyxl
import pandas as pd
import xlrd
from openpyxl.reader.excel import load_workbook
from tkinter import messagebox


class MainWindow:
    def __init__(self, window):
        self.window = window
        self.window.title('Main Window')
        self.widgets = []
        self.create_widgets()

    def create_widgets(self):
        window_width = 350
        window_height = 200
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)
        self.window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        label = tk.Label(text='Choose the program:')
        label.pack(pady=10)
        self.widgets.append(label)

        button2 = tk.Button(text="Ready Excel File For Translation",
                           command=self.switch_to_excel_for_translation)
        button2.pack(pady=10)
        self.widgets.append(button2)

        button = tk.Button(text="Excel Vlookup For Translation",
                           command=self.switch_to_excel_vlookup_for_translation)
        button.pack(pady=10)
        self.widgets.append(button)

    def clear_window(self):
        for widget in self.widgets:
            widget.destroy()
        self.widgets = []

    def switch_to_excel_vlookup_for_translation(self):
        self.clear_window()
        excel_window = ExcelVlookupForTranslation(window)

    def switch_to_excel_for_translation(self):
        self.clear_window()
        excel_window = ReadyExcelFileForTranslation(window)

class ExcelVlookupForTranslation:
    def __init__(self, window):
        self.window = window
        self.window.title('Excel Vlookup For Translation')
        self.widgets = []
        self.create_widgets()

    def create_widgets(self):
        window_width = 350
        window_height = 300
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)
        self.window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        # Create a label and an entry widget for the Excel file
        label1 = tk.Label(text="Choose the Original Excel file:")
        label1.pack()
        self.widgets.append(label1)
        entry1 = tk.Entry(width=45)
        entry1.pack()
        self.widgets.append(entry1)

        # Create a button to open a file dialog
        button1 = tk.Button(text="Select file", command=lambda: self.open_file_dialog(entry1, True))
        button1.pack(pady=10)
        self.widgets.append(button1)

        label2 = tk.Label(text="Choose the translated Excel file:")
        label2.pack()
        self.widgets.append(label2)
        entry2 = tk.Entry(width=45)
        entry2.pack()
        self.widgets.append(entry2)

        # Create a button to open a file dialog
        button2 = tk.Button(text="Select file", command=lambda: self.open_file_dialog(entry2, False))
        button2.pack(pady=10)
        self.widgets.append(button2)

        # Create a button to start the processing
        button3 = tk.Button(text="Process", command=lambda: self.process(entry2))
        button3.pack(pady=10)
        self.widgets.append(button3)

        button4 = tk.Button(text="Back", command=self.switch_to_main_menu)
        button4.pack(pady=10)
        self.widgets.append(button4)

    def clear_window(self):
        for widget in self.widgets:
            widget.destroy()
        self.widgets = []

    def switch_to_main_menu(self):
        self.clear_window()
        main_window = MainWindow(self.window)

    def open_file_dialog(self, entry, original):
        # Open a file dialog and get the selected file's path
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls")])
        if not original:
            global basename_xlxs
            basename_xlxs = os.path.splitext(file_path)[0] + " - Ready For Upload.xlsx"
        if os.path.splitext(file_path)[-1].lower() == ".xls":
            workbook = xlrd.open_workbook(file_path, ignore_workbook_corruption=True)
            # Create a new .xlsx file using openpyxl
            wb = openpyxl.Workbook()

            # Get the sheet names from the .xls file
            sheet_names = workbook.sheet_names()

            # Iterate through the sheet names and copy the data from each sheet
            for sheet_name in sheet_names:
                # Get the sheet from the .xls file
                xls_sheet = workbook.sheet_by_name(sheet_name)

                # Create a new sheet in the .xlsx file
                xlsx_sheet = wb.create_sheet(title=sheet_name)

                # Iterate through the rows and columns of the .xls sheet and copy the data to the .xlsx sheet
                for row in range(xls_sheet.nrows):
                    for col in range(xls_sheet.ncols):
                        xlsx_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

            # Save the .xlsx file
            if original:
                wb.save('otemp.xlsx')
        # Set the file path to the entry widget
        entry.delete(0, tk.END)
        entry.insert(0, file_path)
        if original:
            global ofile
            ofile = 'otemp.xlsx'

    def process(self, entry):
        def process_excel(sheet_name):
            pattern = re.compile('\<(.*?)\>')
            odf = pd.read_excel(ofile, f'{sheet_name}')
            tdf = pd.read_excel(entry.get(), f'{sheet_name}')
            if tdf.columns[0] == 'Object':
                tdf.drop(tdf.columns[0], axis=1, inplace=True)
                odf.drop(odf.columns[[3, 4, 5, 6, 7]], axis=1, inplace=True)
            else:
                odf.drop(odf.columns[[3, 4, 5, 6, 7]], axis=1, inplace=True)
            odf = odf.replace(pattern, '', regex=True)
            df = pd.merge(odf, tdf, left_index=True, right_index=True)
            return df

        wb = load_workbook(entry.get(), read_only=True)

        if 'Settings' in wb:
            global dfsettings
            dfsettings = process_excel("Settings")
        else:
            dfsettings = pd.read_excel(ofile, 'Settings')
        if 'Scales' in wb:
            global dfscales
            dfscales = process_excel("Scales")
        else:
            dfscales = pd.read_excel(ofile, 'Scales')
        if 'Questions' in wb:
            global dfquestions
            dfquestions = process_excel("Questions")
        else:
            dfquestions = pd.read_excel(ofile, 'Questions')

        with pd.ExcelWriter(basename_xlxs) as writer:
            dfsettings.to_excel(writer, sheet_name='Settings', index=False)
            dfscales.to_excel(writer, sheet_name='Scales', index=False)
            dfquestions.to_excel(writer, sheet_name='Questions', index=False)

        os.remove('otemp.xlsx')
        label2 = tk.Label(text="File Saved Successfully")
        label2.pack(pady=10)
        self.widgets.append(label2)

class ReadyExcelFileForTranslation:
    def __init__(self, window):
        self.window = window
        self.window.title('Ready Excel File For Translation')
        self.widgets = []
        self.create_widgets()

    def create_widgets(self):
        window_width = 350
        window_height = 250
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)
        self.window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        # Create a label and an entry widget for the Excel file
        label1 = tk.Label(text="Enter the path of the Excel file:")
        label1.pack(pady=10)
        self.widgets.append(label1)
        entry = tk.Entry(width=45)
        entry.pack(pady=10)
        self.widgets.append(entry)

        # Create a button to open a file dialog
        button1 = tk.Button(text="Select file", command=lambda: self.open_file_dialog(entry))
        button1.pack(pady=10)
        self.widgets.append(button1)

        # Create a button to start the processing
        button2 = tk.Button(text="Process", command=lambda: self.process(entry))
        button2.pack(pady=10)
        self.widgets.append(button2)

        button3 = tk.Button(text="Back", command=self.switch_to_main_menu)
        button3.pack(pady=10)
        self.widgets.append(button3)

    def clear_window(self):
        for widget in self.widgets:
            widget.destroy()
        self.widgets = []

    def switch_to_main_menu(self):
        self.clear_window()
        main_window = MainWindow(self.window)

    def open_file_dialog(self, entry):
        # Open a file dialog and get the selected file's path
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls")])
        global basename_xlxs
        basename_xlxs = os.path.splitext(file_path)[0] + " - Ready For Translation.xlsx"
        if os.path.splitext(file_path)[-1].lower() == ".xls":
            workbook = xlrd.open_workbook(file_path, ignore_workbook_corruption=True)
            # Create a new .xlsx file using openpyxl
            wb = openpyxl.Workbook()

            # Get the sheet names from the .xls file
            sheet_names = workbook.sheet_names()

            # Iterate through the sheet names and copy the data from each sheet
            for sheet_name in sheet_names:
                # Get the sheet from the .xls file
                xls_sheet = workbook.sheet_by_name(sheet_name)

                # Create a new sheet in the .xlsx file
                xlsx_sheet = wb.create_sheet(title=sheet_name)

                # Iterate through the rows and columns of the .xls sheet and copy the data to the .xlsx sheet
                for row in range(xls_sheet.nrows):
                    for col in range(xls_sheet.ncols):
                        xlsx_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

            # Save the .xlsx file
            wb.save('temp.xlsx')
        # Set the file path to the entry widget
        entry.delete(0, tk.END)
        entry.insert(0, file_path)
        global file
        file = 'temp.xlsx'
    def process(self, entry):
        def process_excel(sheet_name):
            pattern = re.compile('\<(.*?)\>')
            df = pd.read_excel(file, f'{sheet_name}')
            df.drop(df.columns[[0, 1, 7]], axis=1, inplace=True)
            df = df.replace(pattern, '', regex=True)
            return df

        dfsettings = process_excel("Settings")
        dfscales = process_excel("Scales")
        dfquestions = process_excel("Questions")

        with pd.ExcelWriter(basename_xlxs) as writer:
            dfsettings.to_excel(writer, sheet_name='Settings', index=False)
            dfscales.to_excel(writer, sheet_name='Scales', index=False)
            dfquestions.to_excel(writer, sheet_name='Questions', index=False)

        os.remove('temp.xlsx')
        label2 = tk.Label(text="File Saved Successfully")
        label2.pack(pady=10)
        self.widgets.append(label2)


if __name__ == '__main__':
    window = tk.Tk()
    main_window = MainWindow(window)
    window.mainloop()
