import os
import re
import tkinter as tk
from tkinter import filedialog

import openpyxl as openpyxl
import pandas as pd
import xlrd as xlrd
from openpyxl.reader.excel import load_workbook


class ExcelVlookupForTranslation:
    def __init__(self, window):
        self.window = window
        self.window.title('Excel Vlookup For Translation')
        self.widgets = []
        self.create_widgets()

    def create_widgets(self):
        window_width = 350
        window_height = 250
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)
        self.window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        # Create a label and an entry widget for the Excel file
        label1 = tk.Label(text="Choose the Original Excel file:")
        label1.pack()
        entry1 = tk.Entry(width=45)
        entry1.pack()

        # Create a button to open a file dialog
        button1 = tk.Button(text="Select file", command=lambda: self.open_file_dialog(entry1, True))
        button1.pack(pady=10)

        label2 = tk.Label(text="Choose the translated Excel file:")
        label2.pack()
        entry2 = tk.Entry(width=45)
        entry2.pack()

        # Create a button to open a file dialog
        button2 = tk.Button(text="Select file", command=lambda: self.open_file_dialog(entry2, False))
        button2.pack(pady=10)

        # Create a button to start the processing
        button3 = tk.Button(text="Process", command=lambda: self.process(entry2))
        button3.pack(pady=10)

    def clear_window(self):
        for widget in self.widgets:
            widget.destroy()
        self.widgets = []

    def switch_to_excel_for_translation(self):
        self.clear_window()

    def open_file_dialog(self, entry, original):
        # Open a file dialog and get the selected file's path
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls")])
        if not original:
            global basename_xlxs
            basename_xlxs = os.path.splitext(file_path)[0] + " - Ready For Upload.xlsx"
        if os.path.splitext(file_path)[-1].lower() == ".xls":
            workbook = xlrd.open_workbook(file_path, ignore_workbook_corruption=True)
            # Create a new .xlsx file using openpyxl
            wb = openpyxl.Workbook()

            # Get the sheet names from the .xls file
            sheet_names = workbook.sheet_names()

            # Iterate through the sheet names and copy the data from each sheet
            for sheet_name in sheet_names:
                # Get the sheet from the .xls file
                xls_sheet = workbook.sheet_by_name(sheet_name)

                # Create a new sheet in the .xlsx file
                xlsx_sheet = wb.create_sheet(title=sheet_name)

                # Iterate through the rows and columns of the .xls sheet and copy the data to the .xlsx sheet
                for row in range(xls_sheet.nrows):
                    for col in range(xls_sheet.ncols):
                        xlsx_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

            # Save the .xlsx file
            if original:
                wb.save('otemp.xlsx')
        # Set the file path to the entry widget
        entry.delete(0, tk.END)
        entry.insert(0, file_path)
        if original:
            global ofile
            ofile = 'otemp.xlsx'

    def process(self, entry):
        def process_excel(sheet_name):
            pattern = re.compile('\<(.*?)\>')
            odf = pd.read_excel(ofile, f'{sheet_name}')
            tdf = pd.read_excel(entry.get(), f'{sheet_name}')
            if tdf.columns[0] == 'Object':
                tdf.drop(tdf.columns[0], axis=1, inplace=True)
                odf.drop(odf.columns[[3, 4, 5, 6, 7]], axis=1, inplace=True)
            else:
                odf.drop(odf.columns[[3, 4, 5, 6, 7]], axis=1, inplace=True)
            odf = odf.replace(pattern, '', regex=True)
            df = pd.merge(odf, tdf, left_index=True, right_index=True)
            return df

        wb = load_workbook(entry.get(), read_only=True)

        if 'Settings' in wb:
            global dfsettings
            dfsettings = process_excel("Settings")
            print('Settings exist')
        else:
            dfsettings = pd.read_excel(ofile, 'Settings')
        if 'Scales' in wb:
            global dfscales
            dfscales = process_excel("Scales")
            print('Scales exist')
        else:
            dfscales = pd.read_excel(ofile, 'Scales')
        if 'Questions' in wb:
            global dfquestions
            dfquestions = process_excel("Questions")
            print('Questions exist')
        else:
            dfquestions = pd.read_excel(ofile, 'Questions')

        with pd.ExcelWriter(basename_xlxs) as writer:
            dfsettings.to_excel(writer, sheet_name='Settings', index=False)
            dfscales.to_excel(writer, sheet_name='Scales', index=False)
            dfquestions.to_excel(writer, sheet_name='Questions', index=False)

        os.remove('otemp.xlsx')
        label2 = tk.Label(text="File Saved Successfully")
        label2.pack(pady=10)
